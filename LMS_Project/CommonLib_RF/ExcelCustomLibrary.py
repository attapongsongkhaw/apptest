import os, sys, warnings, openpyxl, pprint, pyexcel, pyexcel.ext.xls, xlrd
from pyexcel._compact import OrderedDict
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Style

class ExcelCustomLibrary:
    """
    This test library provides keywords to allow opening, reading, writing and saving Excel files.

    *Before running tests in Robot Framework*

    Prior to running tests, ExcelCustomLibrary must first be imported into your Robot test suite.

    Example:
        | Library | ExcelCustomLibrary.py |

    """

    ROBOT_LIBRARY_SCOPE = 'TEST SUITE'
    warnings.filterwarnings("ignore")
    def __init__(self):
        self.wb = None
        self.sheet = None
        self.filename = None
        self.excelDic = None
        self.workingRow = None
        self.columnMapDic = {}
        self.excelHeaderDict = {}
        self.excel_style = {'Normal' : {'FontName':'Tahoma', 'FontSize':10, 'FontColor':'FF000000', 'FontBold':False},
                            'Header' : {'FontName':'Tahoma', 'FontSize':10, 'FontColor':'FFBB00', 'FontBold':True}}

    def open_excel_file(self, filename):
        """
        Opens the Excel file from the path provided in the 'filename' parameter.
        *** Support only XLSX file (Excel 2010 onword)
        
        Arguments:
                |  File Name (string)                      | The file name string value that will be used to open the excel file to perform tests upon.                                  |

        Example:
                | *Keywords*           |  *Parameters*              |
                | Open Excel           |  C:\\Excel.xlsx            |

        """
        self.filename = filename
        if filename.endswith('.xlsx'):
            self.wb = openpyxl.load_workbook(filename)
        elif filename.endswith('.xls'):
            self.open_xls_as_xlsx(filename)
        else:
            print "Not support this extension"

    def open_xls_as_xlsx(self, filename):
        xlsWorkBook = xlrd.open_workbook(filename)
        xlsxWorkBook = openpyxl.Workbook()

        for i in xrange(0, xlsWorkBook.nsheets):
            xlsSheet = xlsWorkBook.sheet_by_index(i)
            sheet = xlsxWorkBook.active if i == 0 else xlsxWorkBook.create_sheet()
            sheet.title = xlsSheet.name

            for row in xrange(0, xlsSheet.nrows):
                for col in xrange(0, xlsSheet.ncols):
                    sheet.cell(row=row+1, column=col+1).value = xlsSheet.cell_value(row, col)

        self.wb = xlsxWorkBook
        
    def select_sheet(self, sheetname):
        self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheetname = sheetname
        self.excelHeaderDict = self.get_excel_column_map_dictionary()

    def get_excel_dictionary(self):
        """
        Returns dictionary of each column of the excel active sheet.

        Example:
                | *Keywords*                        |  *Parameters*                                      |
                | Open Excel                        |  C:\\Excel.xlsx                                    |
                | Get Excel Column Map Dictionary   |                                                    |
        
                ------------------------------------------------------------------------------------------

                Excel.xlsx
                ---------------------------------
                |   Name   |    Age  |  Tel. No |
                ---------------------------------
                |   Name1  |   Age1  |   Tel.1  |
                |   Name2  |   Age2  |   Tel.2  |

                ------------------------------------------------------------------------------------------
                This mehtod will retrun:
                    ([('Name',['Name1', 'Name2']),
                      ('Age',['Age1', 'Age2']),
                      ('Tel No.',['Tel.1', 'Tel.2'])])

	"""
        book = pyexcel.get_book(file_name=self.filename)
        sheet = book[self.sheetname]
        sheet.save_as("temp.xls")
        self.excelDic =  pyexcel.get_dict(file_name="temp.xls", name_columns_by_row=0)
        os.remove("temp.xls")
        return self.excelDic

    def get_column_count(self):
        """
        Returns the number of columns of the excel active sheet.

        Example:
                | *Keywords*          |  *Parameters*               |
                | Open Excel          |  C:\\Excel.xlsx             |
                | Get Column Count    |                             |

        """
        return self.sheet.max_column

    def get_row_count(self):
        """
        Returns the number of rows of the excel active sheet.

        Example:
                | *Keywords*          |  *Parameters*               |
                | Open Excel          |  C:\\Excel.xlsx             |
                | Get Row Count       |                             |

        """
        return self.sheet.max_row

    def save_excel_file(self,filename=""):
        """
        Save currenty workbook open excel workbook

         Example:
                | *Keywords*          |  *Parameters*                                    |
                | Open Excel          |  C:\\Excel.xlsx                                  |
                | Create Column       |  Status                     |    No Run          |
                | Save Excel File     |                                                  |
        
        """
        if not filename:
            self.wb.save(self.filename)
        else:
            self.wb.save(filename)

    def create_column(self,columnName, defaultvalue='N/A'):
        """
        Create new column to the excel active sheet. The new column will be crated after the latest column in the excel active sheet.
        The default value will be updated in every row of new column.
        *** The new column could NOT be created if the column name already exists. 

        Arguments:
                |  Column Name (string)                                | The name of column to be created.  |
                |  Default Value of new Column (string)(default='N/A)  | The default value of the new column   |

        Example:
                | *Keywords*          |  *Parameters*               |
                | Open Excel          |  C:\\Excel.xlsx             |
                | Create Column       |  Status                     |    No Run          |
        
        """
        data = self.get_excel_dictionary()
        if (columnName not in self.excelDic.keys()):
            colCount = self.get_column_count()
            rowCount = self.get_row_count()
            nextColumn = get_column_letter(colCount+1)
            
            # create new column
            headerFont = Font(name=self.excel_style['Header']['FontName'],
                            size=self.excel_style['Header']['FontSize'],
                            color=self.excel_style['Header']['FontColor'],
                            bold=self.excel_style['Header']['FontBold'])
            headerStyle = Style(font=headerFont)
            self.sheet[nextColumn+'1'].style = headerStyle
            self.sheet[nextColumn+'1'] = columnName
            self.sheet.column_dimensions[nextColumn].width = 18
            
            # field the default vaule to the new column
            for row in range(2, rowCount + 1):
                normalFont = Font(name=self.excel_style['Normal']['FontName'],
                            size=self.excel_style['Normal']['FontSize'],
                            color=self.excel_style['Normal']['FontColor'],
                            bold=self.excel_style['Normal']['FontBold'])
                normalStyle = Style(font=normalFont)
                self.sheet[nextColumn+str(row)].style = normalStyle
                self.sheet[nextColumn+str(row)].value = defaultvalue
        
        else:
            print ("Column '" + columnName + "' already exists")
        self.excelHeaderDict = self.get_excel_column_map_dictionary()

    def get_excel_column_map_dictionary(self):
        """
        Returns dictionary of each cell value and column letter of the 1st row from the excel active sheet.

        Example:
                | *Keywords*                        |  *Parameters*                                      |
                | Open Excel                        |  C:\\Excel.xlsx                                    |
                | Get Excel Column Map Dictionary   |                                                    |
        
                ------------------------------------------------------------------------------------------

                Excel.xlsx
                ---------------------------------
                |   Name   |    Age  |  Tel. No |
                ---------------------------------
                |   Name1  |   Age1  |   Tel.1  |
                |   Name2  |   Age2  |   Tel.2  |

                ------------------------------------------------------------------------------------------
                This mehtod will retrun the dictionary:
                    {'Name':'A', 'Age':'B', 'Tel. No':'C'}

        """
        for col in range(1, self.get_column_count() + 1):
            cellCoordinate = get_column_letter(col) + '1'
            cellValue = self.sheet[cellCoordinate].value
            self.columnMapDic.setdefault(cellValue,'')
            self.columnMapDic[cellValue] = get_column_letter(col)
        return self.columnMapDic

    def put_string_to_cell(self, column, row, value):
        """
        Using the excel active sheet, the value of the indicated cell is set to be the string given in the parameter.

        Arguments:
                |  Column (string)        | The column string value that will be used to modify the cell.                                    |
                |  Row (int)           | The row integer value that will be used to modify the cell.                                       |
                |  Value (string)      | The string value that will be added to the specified sheetname at the specified column and row.   |
        Example:

                | *Keywords*           |  *Parameters*                             |
                | Open Excel           |  C:\\Excel.xlsx           |     |         |
                | Put String To Cell   |  A                        |  2  |  Hello  |

        """
        self.sheet[column+str(row)] = value

    def get_string_from_cell(self, column, row):
        return self.sheet[column+str(row)].value

    def set_working_row(self, row):
        self.workingRow = row

    def get_data_from_column(self, columnname):
        columnLetter = self.excelHeaderDict[columnname]
        return self.get_string_from_cell(columnLetter, self.workingRow)

    def put_string_to_column(self, columnname, value):
        columnLetter = self.excelHeaderDict[columnname]
        self.sheet[columnLetter+str(self.workingRow)] = value

    def create_excel_workbook(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def search_row_by_value(self, columnname, value):
        columnLetter = self.excelHeaderDict[columnname]
        for row in range(2, self.get_row_count() + 1):
            cellCoordinate = columnLetter + str(row)
            if self.sheet[cellCoordinate].value == value:
                return row

        return 0 #not found


#x = ExcelCustomLibrary()
#x.open_excel_file("D:\\test.xlsx")
#x.select_sheet("TS01")
#r=x.search_row_by_value("TestCase","TC001")
#print r

#x.set_working_row(3)
#b=x.get_data_from_column("User_ID")
#print b
#b=x.get_data_from_column("New Column")
#print b
#x.put_string_to_column("New Column", "New Data")
#b=x.save_excel_file("D:\\output.xlsx")
#x.close_excel()